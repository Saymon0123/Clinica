"""
Rotas de relatórios e exportações
"""
from fastapi import APIRouter, HTTPException, status, Depends, Response
from fastapi.responses import StreamingResponse
from app.models import ReportRequest
from app.database import supabase
from app.auth import get_current_user, parse_date_br
import json
import io
from datetime import datetime

router = APIRouter()

@router.post("/generate")
async def generate_report(
    report_data: ReportRequest,
    current_user: dict = Depends(get_current_user)
):
    """Gera relatório de agendamentos"""
    try:
        # Construir query
        query = supabase.table("appointments").select("*, doctor_id(full_name, specialty)")
        
        # Funcionário só vê seus dados
        if current_user["role"] != "ADM":
            if report_data.doctor_id and report_data.doctor_id != current_user["id"]:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="Você só pode gerar relatórios dos seus próprios dados"
                )
            query = query.eq("doctor_id", current_user["id"])
        elif report_data.doctor_id:
            query = query.eq("doctor_id", report_data.doctor_id)
        
        # Filtrar por data
        date_from = parse_date_br(report_data.date_from)
        date_to = parse_date_br(report_data.date_to)
        
        query = query.gte("appointment_date", date_from).lte("appointment_date", date_to)
        
        # Filtrar por tipo
        if report_data.appointment_type:
            query = query.eq("appointment_type", report_data.appointment_type.value)
        
        result = query.execute()
        appointments = result.data
        
        # Gerar estatísticas
        total = len(appointments)
        by_status = {}
        by_type = {}
        by_doctor = {}
        
        for apt in appointments:
            status = apt["status"]
            by_status[status] = by_status.get(status, 0) + 1
            
            apt_type = apt["appointment_type"]
            by_type[apt_type] = by_type.get(apt_type, 0) + 1
            
            doctor_name = apt.get("doctor_id", {}).get("full_name", "Desconhecido") if isinstance(apt.get("doctor_id"), dict) else "Desconhecido"
            by_doctor[doctor_name] = by_doctor.get(doctor_name, 0) + 1
        
        report = {
            "period": {
                "from": report_data.date_from,
                "to": report_data.date_to
            },
            "summary": {
                "total_appointments": total,
                "by_status": by_status,
                "by_type": by_type,
                "by_doctor": by_doctor
            },
            "appointments": appointments,
            "generated_at": datetime.now().isoformat(),
            "generated_by": current_user["full_name"]
        }
        
        # Retornar no formato solicitado
        if report_data.export_format == "json":
            return report
        elif report_data.export_format == "csv":
            return export_csv(appointments)
        elif report_data.export_format == "excel":
            return export_excel(appointments)
        elif report_data.export_format == "pdf":
            return export_pdf(report)
        else:
            return report
            
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=str(e)
        )

def export_csv(appointments):
    """Exporta para CSV"""
    import csv
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Cabeçalho
    writer.writerow([
        "ID", "Paciente", "Telefone", "E-mail", "Tipo", "Data", "Hora",
        "Duração (min)", "Status", "Observações"
    ])
    
    # Dados
    for apt in appointments:
        writer.writerow([
            apt["id"],
            apt["patient_name"],
            apt["patient_phone"],
            apt.get("patient_email", ""),
            apt["appointment_type"],
            apt["appointment_date"],
            apt["appointment_time"],
            apt["duration_minutes"],
            apt["status"],
            apt.get("notes", "")
        ])
    
    output.seek(0)
    
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=agendamentos_{datetime.now().strftime('%Y%m%d')}.csv"}
    )

def export_excel(appointments):
    """Exporta para Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Agendamentos"
        
        # Cabeçalho com estilo
        headers = [
            "ID", "Paciente", "Telefone", "E-mail", "Tipo", "Data", "Hora",
            "Duração (min)", "Status", "Observações"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="0ea5e9", end_color="0ea5e9", fill_type="solid")
        
        # Dados
        for row, apt in enumerate(appointments, 2):
            ws.cell(row=row, column=1, value=apt["id"])
            ws.cell(row=row, column=2, value=apt["patient_name"])
            ws.cell(row=row, column=3, value=apt["patient_phone"])
            ws.cell(row=row, column=4, value=apt.get("patient_email", ""))
            ws.cell(row=row, column=5, value=apt["appointment_type"])
            ws.cell(row=row, column=6, value=apt["appointment_date"])
            ws.cell(row=row, column=7, value=apt["appointment_time"])
            ws.cell(row=row, column=8, value=apt["duration_minutes"])
            ws.cell(row=row, column=9, value=apt["status"])
            ws.cell(row=row, column=10, value=apt.get("notes", ""))
        
        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Salvar em memória
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=agendamentos_{datetime.now().strftime('%Y%m%d')}.xlsx"}
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Erro ao gerar Excel: {str(e)}"
        )

def export_pdf(report):
    """Exporta para PDF"""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        elements = []
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#0ea5e9'),
            spaceAfter=30,
            alignment=1  # Center
        )
        
        # Título
        title = Paragraph("Relatório de Agendamentos", title_style)
        elements.append(title)
        elements.append(Spacer(1, 12))
        
        # Período
        period_text = f"Período: {report['period']['from']} a {report['period']['to']}"
        elements.append(Paragraph(period_text, styles['Normal']))
        elements.append(Spacer(1, 12))
        
        # Resumo
        summary = report['summary']
        summary_text = f"""
        <b>Total de Agendamentos:</b> {summary['total_appointments']}<br/>
        <b>Por Status:</b> {', '.join([f'{k}: {v}' for k, v in summary['by_status'].items()])}<br/>
        <b>Por Tipo:</b> {', '.join([f'{k}: {v}' for k, v in summary['by_type'].items()])}
        """
        elements.append(Paragraph(summary_text, styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Tabela de agendamentos
        data = [['Paciente', 'Tipo', 'Data', 'Hora', 'Status']]
        
        for apt in report['appointments'][:50]:  # Limitar a 50 para não ficar muito grande
            data.append([
                apt['patient_name'],
                apt['appointment_type'],
                apt['appointment_date'],
                apt['appointment_time'][:5],  # HH:MM
                apt['status']
            ])
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0ea5e9')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        
        # Gerar PDF
        doc.build(elements)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=agendamentos_{datetime.now().strftime('%Y%m%d')}.pdf"}
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Erro ao gerar PDF: {str(e)}"
        )
